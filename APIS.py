from openpyxl import Workbook
import requests
import json
from datetime import datetime


def scan_call(token, address, offset=10000, start_block = 0, end_block = 99999999, output_dir=''):
    normal_tx = requests.get(f'https://api.{token.api_link}/api?module=account&action=txlist&address={address}&startblock={start_block}&endblock={end_block}&offset={offset}&sort=desc&apikey={token.api_key}')

    if normal_tx.status_code != 200:
        return False
   
    erc20_transfer = requests.get(f'https://api.{token.api_link}/api?module=account&action=tokentx&address={address}&offset={offset}&startblock={start_block}&endblock={end_block}&sort=desc&apikey={token.api_key}')
   
    responses = [normal_tx, erc20_transfer]

    filename = f'{output_dir}/{token.symbol}_{address}.xlsx'
   
    work_book = Workbook()
    work_sheet = work_book.active
    work_sheet.title = "Normal Transactions"
    work_book.create_sheet("ERC20 Transfers")

    for index in range(len(work_book.worksheets)):
        if responses[index].status_code == 200:
            content = json.loads(responses[index].content)
            rows = content['result']
            work_book.active = index
            work_sheet = work_book.active
            work_sheet.append(['Date', 'Block Index', 'Transaction ID', 'Sent', 'Received', 'Asset'])
           
            for row in rows:
                asset = row['tokenSymbol'] if 'tokenSymbol' in row else token.symbol
                tokenMultiplier = 10**int(row['tokenDecimal']) if 'tokenDecimal' in row else token.multiplier
                time = str(datetime.fromtimestamp(int(row['timeStamp'])))
                sent = int(row['value'])/tokenMultiplier if row['from'].lower() == address.lower() else 0
                received = int(row['value'])/tokenMultiplier if row['to'].lower() == address.lower() else 0
               
                work_sheet.append([time, row['blockNumber'], row['hash'], sent, received, asset])

    work_book.save(filename)
   
    return True
   
def chair_call(token, address, output_dir=''):
    txns = requests.get(f'https://api.blockchair.com/{token.api_link}/dashboards/address/{address}?transaction_details=true')
   
    content = json.loads(txns.content)
   
    if content['data'] is None or content['data'][address]['address']['type'] is None:
        return False
       
    filename = f'{output_dir}/{token.symbol}_{address}.xlsx'
   
    work_book = Workbook()
    work_sheet = work_book.active
    work_sheet.append(['Date', 'Block Index', 'Transaction ID', 'Sent', 'Received', 'Asset'])
   
    rows = content['data'][address]['transactions']
   
    for row in rows:
        value = int(row['balance_change'])
        sent = value*-1/token.multiplier if value < 0 else 0
        received = value/token.multiplier if value > 0 else 0
       
        work_sheet.append([row['time'], row['block_id'], row['hash'], sent, received, token.symbol])        
   
    work_book.save(filename)
   
    return True
   
   
def algo_call(token, address, output_dir=''):
    txns = requests.get(f'https://algoindexer.algoexplorerapi.io/v2/transactions?address={address}')
   
    content = json.loads(txns.content)
   
    if 'message' in content:
        return False

    filename = f'{output_dir}/{token.symbol}_{address}.xlsx'
   
    work_book = Workbook()
    work_sheet = work_book.active
    work_sheet.append(['Date', 'Block Index', 'Transaction ID', 'Sent', 'Received', 'Asset'])
   
    rows = content['transactions']
   
    for row in rows:
        time = str(datetime.fromtimestamp(int(row['round-time'])))
        value = int(row['payment-transaction']['amount'])
        sent = value/token.multiplier if row['payment-transaction']['receiver'].lower() != address.lower() else 0
        received = value/token.multiplier if row['payment-transaction']['receiver'].lower() == address.lower() else 0
       
        work_sheet.append([time, row['confirmed-round'], row['id'], sent, received, token.symbol])        
   
    work_book.save(filename)    
   
    return True 